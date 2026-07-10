import argparse
import re
from pathlib import Path
from typing import BinaryIO, Iterable, Optional, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = [
    "ac",
    "wo",
    "eo",
    "eo_description",
]

OUTPUT_COLUMNS = [
    "AC",
    "WO",
    "EO",
    "DESCRIPTION",
    "P/N",
    "S/N",
    "CATEGORY",
    "MODIFIED_DATE",
    "AC_TYPE",
    "AC_SERIES",
    "REMAINING_HOURS",
    "REMAINING_MINUTES",
    "REMAINING_CYCLES",
    "REMAINING_DAYS",
    "COMP_HOURS",
    "TOT_HOURS",
    "TOT_CYCLES",
    "CONTROL",
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

GROUP_FILL = PatternFill("solid", fgColor="D9EAF7")
CHECK_FILL = PatternFill("solid", fgColor="E2F0D9")
DEFECT_FILL = PatternFill("solid", fgColor="FCE4D6")

THIN_GRAY = Side(style="thin", color="D9E2F3")
MEDIUM_BLUE = Side(style="medium", color="5B9BD5")


def clean_col_name(value):
    return str(value).strip().lower().replace(" ", "_")


def is_blank(value):
    if pd.isna(value):
        return True

    if str(value).strip() == "":
        return True

    return False


def text(value):
    if is_blank(value):
        return ""

    return str(value).strip()


def normalize_task(value):
    clean_value = text(value).upper()
    clean_value = re.sub(r"\s+", " ", clean_value)
    return clean_value


def read_program_file(uploaded_file, filename=None):
    name = filename or getattr(uploaded_file, "name", "") or str(uploaded_file)
    suffix = Path(name).suffix.lower()

    if suffix == ".xls":
        engine = "xlrd"
    elif suffix == ".xlsx":
        engine = "openpyxl"
    else:
        raise ValueError("Formato no soportado. Sube un archivo .xls o .xlsx.")

    df = pd.read_excel(uploaded_file, engine=engine)

    df.columns = [clean_col_name(column) for column in df.columns]

    missing_columns = []

    for column in REQUIRED_COLUMNS:
        if column not in df.columns:
            missing_columns.append(column)

    if missing_columns:
        raise ValueError(
            "Faltan columnas requeridas: " + ", ".join(missing_columns)
        )

    return df


def build_full_description(row):
    description = text(row.get("eo_description"))
    pn = text(row.get("pn"))
    sn = text(row.get("sn"))

    if pn and sn:
        pn_sn_text = "P/N: " + pn + " | S/N: " + sn

        if description:
            description = description + "\n" + pn_sn_text
        else:
            description = pn_sn_text

    return description


def is_defect(row):
    control = normalize_task(row.get("control"))

    defect_columns = [
        "wo_task_card_defect",
        "wo_task_card_non_routine",
        "wo_task_card_defect_type",
        "wo_task_card_defect_item",
        "wo_task_card_task_card",
        "wo_task_card_description",
    ]

    has_defect_data = False

    for column in defect_columns:
        if column in row.index:
            if not is_blank(row.get(column)):
                has_defect_data = True
                break

    if control == "Y":
        return True

    if has_defect_data:
        return True

    return False


def task_category(row):
    eo = normalize_task(row.get("eo"))
    description = normalize_task(row.get("eo_description"))

    if is_defect(row):
        return 4, "DEFECT"

    if eo == "WEEKLY CHECK" or description == "WEEKLY CHECK":
        return 1, "WEEKLY CHECK"

    if eo == "DAILY CHECK" or description == "DAILY CHECK":
        return 2, "DAILY CHECK"

    return 3, "EO"


def transform(df):
    work = df.copy()

    work["__original_order"] = range(len(work))

    for column in ["ac", "wo", "eo", "pn", "sn"]:
        if column not in work.columns:
            work[column] = ""

    category_results = work.apply(task_category, axis=1, result_type="expand")

    work["__category_order"] = category_results[0]
    work["category"] = category_results[1]

    work["description_full"] = work.apply(build_full_description, axis=1)

    work = work.sort_values(
        by=[
            "ac",
            "wo",
            "__category_order",
            "__original_order",
        ],
        kind="mergesort",
        na_position="last",
    )

    output = pd.DataFrame(
        {
            "AC": work["ac"],
            "WO": work["wo"],
            "EO": work["eo"],
            "DESCRIPTION": work["description_full"],
            "P/N": work["pn"],
            "S/N": work["sn"],
            "CATEGORY": work["category"],
            "MODIFIED_DATE": work["modified_date"] if "modified_date" in work.columns else "",
            "AC_TYPE": work["ac_type"] if "ac_type" in work.columns else "",
            "AC_SERIES": work["ac_series"] if "ac_series" in work.columns else "",
            "REMAINING_HOURS": work["remaining_hours"] if "remaining_hours" in work.columns else "",
            "REMAINING_MINUTES": work["remaining_minutes"] if "remaining_minutes" in work.columns else "",
            "REMAINING_CYCLES": work["remaining_cycles"] if "remaining_cycles" in work.columns else "",
            "REMAINING_DAYS": work["remaining_days"] if "remaining_days" in work.columns else "",
            "COMP_HOURS": work["comp_hours"] if "comp_hours" in work.columns else "",
            "TOT_HOURS": work["tot_hours"] if "tot_hours" in work.columns else "",
            "TOT_CYCLES": work["tot_cycles"] if "tot_cycles" in work.columns else "",
            "CONTROL": work["control"] if "control" in work.columns else "",
        }
    )

    output = output[OUTPUT_COLUMNS]

    return output


def best_width(values, min_width, max_width):
    width = min_width

    for value in values:
        clean_value = text(value).replace("\n", " ")
        calculated_width = len(clean_value) + 2

        if calculated_width > width:
            width = calculated_width

        if width > max_width:
            width = max_width

    return width


def apply_header_style(ws):
    for cell in wscell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=MEDIUM_BLUE)


def apply_body_style(ws, max_row, max_col):
    for row in ws.iter_rows(
        min_row=2,
        max_row=max_row,
        max_col=max_col,
    ):
        category = text(row[6].value)
        fill = None

        if category in ["WEEKLY CHECK", "DAILY CHECK"]:
            fill = CHECK_FILL

        if category == "DEFECT":
            fill = DEFECT_FILL

        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = Border(bottom=THIN_GRAY)

            if fill is not None:
                cell.fill = fill

        row[3].alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        ws.row_dimensions[row[0].row].height = 42


def merge_ac_wo_groups(ws, max_row, max_col):
    group_start = 2

    while group_start <= max_row:
        ac = ws.cell(group_start, 1).value
        wo = ws.cell(group_start, 2).value

        group_end = group_start

        while group_end + 1 <= max_row:
            next_ac = ws.cell(group_end + 1, 1).value
            next_wo = ws.cell(group_end + 1, 2).value

            if next_ac == ac and next_wo == wo:
                group_end = group_end + 1
            else:
                break

        if group_end > group_start:
            ws.merge_cells(
                start_row=group_start,
                start_column=1,
                end_row=group_end,
                end_column=1,
            )

            ws.merge_cells(
                start_row=group_start,
                start_column=2,
                end_row=group_end,
                end_column=2,
            )

        for column in [1, 2]:
            cell = ws.cell(group_start, column)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.fill = GROUP_FILL
            cell.font = Font(bold=True)

        for column in range(1, max_col + 1):
            ws.cell(group_start, column).border = Border(
                top=MEDIUM_BLUE,
                bottom=THIN_GRAY,
            )

        group_start = group_end + 1


def apply_column_widths(ws, df_out):
    fixed_widths = {
        "A": 12,
        "B": 12,
        "C": 18,
        "D": 62,
        "E": 18,
        "F": 18,
        "G": 16,
        "H": 20,
        "I": 12,
        "J": 12,
        "K": 16,
        "L": 18,
        "M": 17,
        "N": 16,
        "O": 13,
        "P": 13,
        "Q": 12,
        "R": 10,
    }

    for column_index, header in enumerate(OUTPUT_COLUMNS, start=1):
        letter = get_column_letter(column_index)

        if letter in fixed_widths:
            ws.column_dimensions[letter].width = fixed_widths[letter]
        else:
            ws.column_dimensions[letter].width = best_width(
                [header] + list(df_out[header]),
                10,
                30,
            )


def write_formatted_excel(df_out, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Programa Ordenado"

    ws.append(OUTPUT_COLUMNS)

    for row in df_out.itertuples(index=False, name=None):
        ws.append(list(row))

    max_row = ws.max_row
    max_col = ws.max_column

    apply_header_style(ws)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False

    apply_body_style(
        ws=ws,
        max_row=max_row,
        max_col=max_col,
    )

    merge_ac_wo_groups(
        ws=ws,
        max_row=max_row,
        max_col=max_col,
    )

    apply_column_widths(
        ws=ws,
        df_out=df_out,
    )

    for cell in ws["H"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm"

    ws.auto_filter.ref = "C1:" + get_column_letter(max_col) + str(max_row)

    output_path = Path(output_path)
    wb.save(output_path)

    return output_path


def process_file(input_file, output_path, filename=None):
    df = read_program_file(
        uploaded_file=input_file,
        filename=filename,
    )

    df_out = transform(df)

    return write_formatted_excel(
        df_out=df_out,
        output_path=output_path,
    )


def main():
    parser = argparse.ArgumentParser(
        description="Ordena programa de mantenimiento por AC/WO y tipo de tarea."
    )

    parser.add_argument(
        "input",
        help="Archivo fuente .xls o .xlsx",
    )

    parser.add_argument(
        "output",
        nargs="?",
        help="Archivo salida .xlsx",
    )

    args = parser.parse_args()

    input_path = Path(args.input)

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = input_path.with_name(input_path.stem + "_ordenado.xlsx")

    process_file(
        input_file=input_path,
        output_path=output_path,
    )

    print("Archivo generado:", output_path)


if __name__ == "__main__":
    main()

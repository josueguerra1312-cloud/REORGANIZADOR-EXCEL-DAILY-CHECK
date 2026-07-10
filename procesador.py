from io import BytesIO
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLUMNAS_SALIDA = ["AC", "WO", "TASK", "DESCRIPTION"]


def limpiar_texto(valor):
    """
    Limpia valores provenientes de Excel.
    Convierte NaN, None o espacios vacios en cadena vacia.
    Conserva saltos de linea internos en descripciones.
    """
    if pd.isna(valor):
        return ""

    texto = str(valor).replace("\r\n", "\n").replace("\r", "\n")
    lineas = [linea.rstrip() for linea in texto.split("\n")]
    texto = "\n".join(lineas).strip()

    if texto.lower() in ["nan", "none", "nat"]:
        return ""

    return texto


def normalizar_columnas(df):
    """
    Normaliza nombres de columnas para trabajar sin depender de mayusculas,
    espacios o pequenas variaciones.
    """
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(" ", "_", regex=False)
    )
    return df


def validar_columnas(df):
    columnas_requeridas = ["ac", "wo", "eo", "eo_description"]
    faltantes = [col for col in columnas_requeridas if col not in df.columns]

    if faltantes:
        raise ValueError(
            "El archivo no contiene las columnas requeridas: "
            + ", ".join(faltantes)
            + ". Revisa que el Excel tenga al menos: ac, wo, eo y eo_description."
        )


def construir_descripcion(row):
    """
    Usa siempre el texto completo de eo_description.
    Agrega P/N y S/N solo si ambos estan presentes.
    """
    descripcion = limpiar_texto(row.get("eo_description", ""))
    pn = limpiar_texto(row.get("pn", ""))
    sn = limpiar_texto(row.get("sn", ""))

    if pn and sn:
        texto_pn_sn = f"P/N {pn} S/N {sn}"
        if descripcion:
            descripcion = f"{descripcion}\n{texto_pn_sn}"
        else:
            descripcion = texto_pn_sn

    return descripcion


def es_defecto(row):
    """
    Identifica defectos para enviarlos al final del grupo.

    En el archivo de entrada se observa columna control con valores como:
    - N: tarea normal
    - P: tarea con P/N y S/N, pero no necesariamente defecto
    - Y: item marcado como defecto/componente asociado

    Tambien se consideran columnas de defecto si vienen pobladas.
    """
    control = limpiar_texto(row.get("control", "")).upper()
    defect = limpiar_texto(row.get("wo_task_card_defect", "")).upper()
    non_routine = limpiar_texto(row.get("wo_task_card_non_routine", "")).upper()
    defect_type = limpiar_texto(row.get("wo_task_card_defect_type", ""))

    if control == "Y":
        return True
    if defect in ["Y", "YES", "TRUE", "1"]:
        return True
    if non_routine in ["Y", "YES", "TRUE", "1"]:
        return True
    if defect_type:
        return True

    return False


def prioridad_tarea(row):
    """
    Orden solicitado dentro de cada grupo AC + WO:
    1. WEEKLY CHECK
    2. DAILY CHECK
    3. EO normales
    4. Defectos
    """
    task = limpiar_texto(row.get("TASK", "")).upper()

    if task == "WEEKLY CHECK":
        return 0
    if task == "DAILY CHECK":
        return 1
    if row.get("_ES_DEFECTO", False):
        return 3

    return 2


def procesar_dataframe(df):
    """
    Recibe un DataFrame del Excel original y genera el DataFrame final ordenado.
    """
    df = normalizar_columnas(df)
    validar_columnas(df)

    df = df.copy()
    df["_orden_original"] = range(len(df))

    filas = []

    for _, row in df.iterrows():
        ac = limpiar_texto(row.get("ac", ""))
        wo = limpiar_texto(row.get("wo", ""))
        task = limpiar_texto(row.get("eo", ""))

        if not ac or not wo or not task:
            continue

        descripcion = construir_descripcion(row)
        defecto = es_defecto(row)

        filas.append(
            {
                "AC": ac,
                "WO": wo,
                "TASK": task,
                "DESCRIPTION": descripcion,
                "_ES_DEFECTO": defecto,
                "_ORDEN_ORIGINAL": row["_orden_original"],
            }
        )

    if not filas:
        raise ValueError("No se encontraron tareas validas para procesar.")

    salida = pd.DataFrame(filas)
    salida["_PRIORIDAD"] = salida.apply(prioridad_tarea, axis=1)

    salida = salida.sort_values(
        by=["AC", "WO", "_PRIORIDAD", "TASK", "_ORDEN_ORIGINAL"],
        ascending=[True, True, True, True, True],
        kind="mergesort",
    )

    salida = salida[COLUMNAS_SALIDA]
    return salida.reset_index(drop=True)


def leer_excel(archivo):
    """
    Lee archivos .xls o .xlsx desde ruta local o desde Streamlit UploadedFile.
    """
    nombre = getattr(archivo, "name", None)

    if nombre:
        extension = Path(nombre).suffix.lower()
    else:
        extension = Path(str(archivo)).suffix.lower()

    if extension == ".xls":
        engine = "xlrd"
    elif extension == ".xlsx":
        engine = "openpyxl"
    else:
        raise ValueError("Formato no soportado. Usa un archivo .xls o .xlsx.")

    return pd.read_excel(archivo, dtype=str, engine=engine)


def aplicar_formato_excel(ws):
    """
    Aplica formato profesional:
    - Encabezados destacados
    - Anchos ajustados
    - Texto con ajuste
    - AC y WO combinados por grupo
    """
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    thin_gray = Side(style="thin", color="D9E2F3")
    border_bottom = Border(bottom=thin_gray)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_bottom

    widths = {
        "A": 14,
        "B": 14,
        "C": 22,
        "D": 95,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(
                vertical="top",
                horizontal="left",
                wrap_text=True,
            )
            cell.border = border_bottom

    for row_idx in range(2, ws.max_row + 1):
        descripcion = ws.cell(row=row_idx, column=4).value or ""
        saltos = str(descripcion).count("\n") + 1
        ws.row_dimensions[row_idx].height = max(24, min(120, saltos * 16))

    # Combinar AC y WO por grupos consecutivos
    if ws.max_row < 2:
        return

    inicio_grupo = 2
    ac_actual = ws.cell(row=2, column=1).value
    wo_actual = ws.cell(row=2, column=2).value

    for row_idx in range(3, ws.max_row + 2):
        if row_idx <= ws.max_row:
            ac = ws.cell(row=row_idx, column=1).value
            wo = ws.cell(row=row_idx, column=2).value
        else:
            ac = None
            wo = None

        cambio_grupo = ac != ac_actual or wo != wo_actual

        if cambio_grupo:
            fin_grupo = row_idx - 1

            if fin_grupo > inicio_grupo:
                ws.merge_cells(
                    start_row=inicio_grupo,
                    start_column=1,
                    end_row=fin_grupo,
                    end_column=1,
                )
                ws.merge_cells(
                    start_row=inicio_grupo,
                    start_column=2,
                    end_row=fin_grupo,
                    end_column=2,
                )

                ws.cell(row=inicio_grupo, column=1).alignment = Alignment(
                    vertical="top",
                    horizontal="center",
                    wrap_text=True,
                )
                ws.cell(row=inicio_grupo, column=2).alignment = Alignment(
                    vertical="top",
                    horizontal="center",
                    wrap_text=True,
                )

            inicio_grupo = row_idx
            ac_actual = ac
            wo_actual = wo

    # Reafirma estilo de fuente general
    for col_idx in range(1, ws.max_column + 1):
        for cell in ws[get_column_letter(col_idx)]:
            if cell.row != 1:
                cell.font = Font(name="Calibri", size=11)


def dataframe_a_excel(df):
    """
    Convierte el DataFrame final en archivo Excel en memoria.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Programa"

    ws.append(COLUMNAS_SALIDA)

    for _, row in df.iterrows():
        ws.append(
            [
                row["AC"],
                row["WO"],
                row["TASK"],
                row["DESCRIPTION"],
            ]
        )

    aplicar_formato_excel(ws)

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida


def procesar_excel(archivo):
    """
    Funcion principal para Streamlit.
    Recibe archivo cargado y devuelve BytesIO con el Excel procesado.
    """
    df = leer_excel(archivo)
    df_final = procesar_dataframe(df)
    return dataframe_a_excel(df_final)


def procesar_y_guardar(ruta_entrada, ruta_salida):
    """
    Funcion para uso local por terminal:
    python procesador.py entrada.xls salida.xlsx
    """
    df = leer_excel(ruta_entrada)
    df_final = procesar_dataframe(df)
    archivo_excel = dataframe_a_excel(df_final)

    with open(ruta_salida, "wb") as f:
        f.write(archivo_excel.getvalue())


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Uso:")
        print("python procesador.py archivo_entrada.xls archivo_salida.xlsx")
        sys.exit(1)

    entrada = sys.argv[1]
    salida = sys.argv[2]

    procesar_y_guardar(entrada, salida)
    print(f"Archivo generado correctamente: {salida}")

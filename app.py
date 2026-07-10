import io
import re
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = ["ac", "wo", "eo", "eo_description"]

OUTPUT_COLUMNS = [
    "AC",
    "WO",
    "EO",
    "DESCRIPTION",
    "P/N",
    "S/N",
    "CATEGORY",
    "MODIFIED_DATE",
    "AC_TYPE",
    "AC_SERIES",
    "REMAINING_HOURS",
    "REMAINING_MINUTES",
    "REMAINING_CYCLES",
    "REMAINING_DAYS",
    "COMP_HOURS",
    "TOT_HOURS",
    "TOT_CYCLES",
    "CONTROL",
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

GROUP_FILL = PatternFill("solid", fgColor="D9EAF7")
CHECK_FILL = PatternFill("solid", fgColor="E2F0D9")
DEFECT_FILL = PatternFill("solid", fgColor="FCE4D6")

THIN_GRAY = Side(style="thin", color="D9E2F3")
MEDIUM_BLUE = Side(style="medium", color="5B9BD5")


def clean_col_name(value):
    return str(value).strip().lower().replace(" ", "_")


def is_blank(value):
    if pd.isna(value):
        return True
    return str(value).strip() == ""


def text(value):
    if is_blank(value):
        return ""
    return str(value).strip()


def normalize_task(value):
    return re.sub(r"\s+", " ", text(value).upper())


def read_excel_file(uploaded_file):
    suffix = Path(uploaded_file.name).suffix.lower()

    uploaded_file.seek(0)

    if suffix == ".xls":
        df = pd.read_excel(uploaded_file, engine="xlrd")
    elif suffix == ".xlsx":
        df = pd.read_excel(uploaded_file, engine="openpyxl")
    else:
        raise ValueError("Formato no soportado. Usa un archivo .xls o .xlsx.")

    df.columns = [clean_col_name(column) for column in df.columns]

    missing_columns = [
        column for column in REQUIRED_COLUMNS if column not in df.columns
    ]

    if missing_columns:
        raise ValueError(
            "Faltan columnas requeridas: " + ", ".join(missing_columns)
        )

    return df


def is_defect(row):
    control = normalize_task(row.get("control"))

    defect_columns = [
        "wo_task_card_defect",
        "wo_task_card_non_routine",
        "wo_task_card_defect_type",
        "wo_task_card_defect_item",
        "wo_task_card_task_card",
        "wo_task_card_description",
    ]

    if control == "Y":
        return True

    for column in defect_columns:
        if column in row.index and not is_blank(row.get(column)):
            return True

    return False


def task_category(row):
    eo = normalize_task(row.get("eo"))
    description = normalize_task(row.get("eo_description"))

    if is_defect(row):
        return 4, "DEFECT"

    if eo == "WEEKLY CHECK" or description == "WEEKLY CHECK":
        return 1, "WEEKLY CHECK"

    if eo == "DAILY CHECK" or description == "DAILY CHECK":
        return 2, "DAILY CHECK"

    return 3, "EO"


def build_description(row):
    description = text(row.get("eo_description"))
    pn = text(row.get("pn"))
    sn = text(row.get("sn"))

    if pn and sn:
        pn_sn = "P/N: " + pn + " | S/N: " + sn

        if description:
            description = description + "\n" + pn_sn
        else:
            description = pn_sn

    return description


def column_values(df, column_name):
    if column_name in df.columns:
        return df[column_name].tolist()

    return [""] * len(df)


def transform_dataframe(df):
    work = df.copy()
    work["__original_order"] = range(len(work))

    for column in ["ac", "wo", "eo", "pn", "sn"]:
        if column not in work.columns:
            work[column] = ""

    category_data = work.apply(task_category, axis=1, result_type="expand")

    work["__category_order"] = category_data[0]
    work["category"] = category_data[1]
    work["description_full"] = work.apply(build_description, axis=1)

    work = work.sort_values(
        by=["ac", "wo", "__category_order", "__original_order"],
        kind="mergesort",
        na_position="last",
    ).reset_index(drop=True)

    output = pd.DataFrame(
        {
            "AC": column_values(work, "ac"),
            "WO": column_values(work, "wo"),
            "EO": column_values(work, "eo"),
            "DESCRIPTION": column_values(work, "description_full"),
            "P/N": column_values(work, "pn"),
            "S/N": column_values(work, "sn"),
            "CATEGORY": column_values(work, "category"),
            "MODIFIED_DATE": column_values(work, "modified_date"),
            "AC_TYPE": column_values(work, "ac_type"),
            "AC_SERIES": column_values(work, "ac_series"),
            "REMAINING_HOURS": column_values(work, "remaining_hours"),
            "REMAINING_MINUTES": column_values(work, "remaining_minutes"),
            "REMAINING_CYCLES": column_values(work, "remaining_cycles"),
            "REMAINING_DAYS": column_values(work, "remaining_days"),
            "COMP_HOURS": column_values(work, "comp_hours"),
            "TOT_HOURS": column_values(work, "tot_hours"),
            "TOT_CYCLES": column_values(work, "tot_cycles"),
            "CONTROL": column_values(work, "control"),
        }
    )

    return output[OUTPUT_COLUMNS]


def apply_header_style(ws):
    for cell in wscell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=MEDIUM_BLUE)


def apply_body_style(ws, max_row, max_col):
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        category = text(row[6].value)
        fill = None

        if category in ["WEEKLY CHECK", "DAILY CHECK"]:
            fill = CHECK_FILL
        elif category == "DEFECT":
            fill = DEFECT_FILL

        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = Border(bottom=THIN_GRAY)

            if fill is not None:
                cell.fill = fill

        row[3].alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        ws.row_dimensions[row[0].row].height = 42


def merge_ac_wo_groups(ws, max_row, max_col):
    group_start = 2

    while group_start <= max_row:
        ac = ws.cell(group_start, 1).value
        wo = ws.cell(group_start, 2).value

        group_end = group_start

        while group_end + 1 <= max_row:
            next_ac = ws.cell(group_end + 1, 1).value
            next_wo = ws.cell(group_end + 1, 2).value

            if next_ac == ac and next_wo == wo:
                group_end += 1
            else:
                break

        if group_end > group_start:
            ws.merge_cells(
                start_row=group_start,
                start_column=1,
                end_row=group_end,
                end_column=1,
            )

            ws.merge_cells(
                start_row=group_start,
                start_column=2,
                end_row=group_end,
                end_column=2,
            )

        for column in [1, 2]:
            cell = ws.cell(group_start, column)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.fill = GROUP_FILL
            cell.font = Font(bold=True)

        for column in range(1, max_col + 1):
            ws.cell(group_start, column).border = Border(
                top=MEDIUM_BLUE,
                bottom=THIN_GRAY,
            )

        group_start = group_end + 1


def apply_column_widths(ws):
    widths = {
        "A": 12,
        "B": 12,
        "C": 18,
        "D": 62,
        "E": 18,
        "F": 18,
        "G": 16,
        "H": 20,
        "I": 12,
        "J": 12,
        "K": 16,
        "L": 18,
        "M": 17,
        "N": 16,
        "O": 13,
        "P": 13,
        "Q": 12,
        "R": 10,
    }

    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def create_excel_bytes(df_out):
    wb = Workbook()
    ws = wb.active
    ws.title = "Programa Ordenado"

    ws.append(OUTPUT_COLUMNS)

    for row in df_out.itertuples(index=False, name=None):
        ws.append(list(row))

    max_row = ws.max_row
    max_col = ws.max_column

    apply_header_style(ws)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False

    apply_body_style(ws, max_row, max_col)
    merge_ac_wo_groups(ws, max_row, max_col)
    apply_column_widths(ws)

    for cell in ws["H"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm"

    ws.auto_filter.ref = "C1:" + get_column_letter(max_col) + str(max_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


st.set_page_config(
    page_title="Reorganizador Excel Daily Check",
    page_icon="✈️",
    layout="centered",
)


st.title("✈️ Reorganizador Excel Daily Check")

st.write(
    "Carga tu archivo Excel y descarga una versión ordenada por AC y WO."
)

st.markdown(
    """
**Orden dentro de cada grupo AC/WO:**

1. WEEKLY CHECK  
2. DAILY CHECK  
3. EO normales  
4. Defectos / Non-Routines
"""
)


uploaded_file = st.file_uploader(
    "Sube tu archivo .xls o .xlsx",
    type=["xls", "xlsx"],
)


if uploaded_file is not None:
    st.success("Archivo cargado: " + uploaded_file.name)

    if st.button("Generar Excel ordenado", type="primary"):
        try:
            with st.spinner("Procesando..."):
                source_df = read_excel_file(uploaded_file)
                ordered_df = transform_dataframe(source_df)
                excel_bytes = create_excel_bytes(ordered_df)
                output_name = Path(uploaded_file.name).stem + "_ordenado.xlsx"

            st.success("Archivo generado correctamente.")

            st.download_button(
                label="Descargar Excel ordenado",
                data=excel_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as error:
            st.error("No se pudo procesar el archivo.")
            st.exception(error)

else:
    st.info("Sube un archivo para comenzar.")

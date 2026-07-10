from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_COLUMNS = ["AC", "WO", "TASK", "DESCRIPTION"]


def _norm_col(col: str) -> str:
    """Normaliza nombres de columnas para tolerar mayusculas, espacios y acentos."""
    s = str(col).strip().lower()
    s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def _clean(value) -> str:
    """Convierte NaN/None/espacios raros a texto limpio, conservando saltos de linea."""
    if pd.isna(value):
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    return text.strip()


def _is_blank(value) -> bool:
    return _clean(value) == ""


def _is_yes(value) -> bool:
    return _clean(value).upper() in {"Y", "YES", "SI", "S", "TRUE", "1"}


def _task_rank(row: pd.Series) -> int:
    """Orden solicitado: WEEKLY, DAILY, EO normales, defectos."""
    task = _clean(row.get("_task", "")).upper()
    is_defect = bool(row.get("_is_defect", False))

    if task == "WEEKLY CHECK":
        return 0
    if task == "DAILY CHECK":
        return 1
    if is_defect:
        return 3
    return 2


def _build_task(row: pd.Series) -> str:
    """Usa EO; si viene vacio, intenta usar campos de task card/defecto."""
    for col in ["eo", "wo_task_card_task_card", "wo_task_card_defect_type", "wo_task_card_defect_item"]:
        value = _clean(row.get(col, ""))
        if value:
            return value
    return "SIN TASK"


def _build_description(row: pd.Series) -> str:
    """
    Descripcion = EO_DESCRIPTION completo.
    Si PN y SN ambos existen, agrega: P/N xxx S/N yyy.
    Si EO_DESCRIPTION viene vacio, usa descripcion/notas de defectos como respaldo.
    """
    desc = _clean(row.get("eo_description", ""))

    if not desc:
        # Respaldo para defectos o task cards sin EO_DESCRIPTION.
        candidates = [
            _clean(row.get("wo_task_card_description", "")),
            _clean(row.get("wo_task_card_notes", "")),
            _clean(row.get("eo_notes", "")),
        ]
        desc = "\n".join([c for c in candidates if c])

    pn = _clean(row.get("pn", ""))
    sn = _clean(row.get("sn", ""))
    if pn and sn:
        suffix = f"P/N {pn} S/N {sn}"
        desc = f"{desc} {suffix}".strip() if desc else suffix

    return desc


def leer_excel_entrada(file) -> pd.DataFrame:
    """
    Lee .xls o .xlsx. En Streamlit, file puede ser UploadedFile/BytesIO.
    En uso local, file puede ser una ruta.
    """
    name = getattr(file, "name", str(file)).lower()
    engine = "xlrd" if name.endswith(".xls") else "openpyxl"
    return pd.read_excel(file, engine=engine, dtype=object)


def procesar_programa(
    input_file,
    incluir_defectos_sin_eo: bool = False,
    ordenar_grupos_por: str = "ac_wo",
) -> pd.DataFrame:
    """
    Transforma el archivo original al formato final:
    AC | WO | TASK | DESCRIPTION.

    Parametros:
    - incluir_defectos_sin_eo=False: replica el comportamiento comun del ejemplo,
      conservando defectos con EO/control, pero omitiendo N/R sin EO si se desea.
    - ordenar_grupos_por='ac_wo' o 'aparicion'.
    """
    df = leer_excel_entrada(input_file)
    df.columns = [_norm_col(c) for c in df.columns]

    required = {"ac", "wo", "eo", "eo_description"}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"Faltan columnas requeridas en el archivo: {', '.join(missing)}")

    # Asegurar columnas opcionales.
    optional_cols = [
        "pn", "sn", "eo_notes", "wo_task_card_task_card", "wo_task_card_description",
        "wo_task_card_non_routine", "wo_task_card_defect", "wo_task_card_defect_type",
        "wo_task_card_defect_item", "wo_task_card_notes", "control",
    ]
    for col in optional_cols:
        if col not in df.columns:
            df[col] = ""

    df = df.copy()
    df["_original_order"] = range(len(df))
    df["_task"] = df.apply(_build_task, axis=1)
    df["_description"] = df.apply(_build_description, axis=1)
    df["_is_defect"] = df.apply(
        lambda r: _is_yes(r.get("wo_task_card_defect", "")) or _is_yes(r.get("wo_task_card_non_routine", "")) or _is_yes(r.get("control", "")),
        axis=1,
    )
    df["_rank"] = df.apply(_task_rank, axis=1)

    # Quitar renglones completamente inutiles.
    df = df[~(df["ac"].apply(_is_blank) & df["wo"].apply(_is_blank) & df["_task"].eq("SIN TASK"))]

    # Para parecerse al archivo ejemplo: omite defectos N/R sin EO si no se pide incluirlos.
    if not incluir_defectos_sin_eo:
        df = df[~(df["eo"].apply(_is_blank) & df["_is_defect"])]

    if ordenar_grupos_por == "aparicion":
        group_sequence = df.groupby(["ac", "wo"], dropna=False)["_original_order"].transform("min")
        df["_group_order"] = group_sequence
        sort_cols = ["_group_order", "_rank", "_original_order"]
    else:
        sort_cols = ["ac", "wo", "_rank", "_original_order"]

    df = df.sort_values(sort_cols, kind="mergesort")

    out = pd.DataFrame({
        "AC": df["ac"].apply(_clean),
        "WO": df["wo"].apply(_clean),
        "TASK": df["_task"].apply(_clean),
        "DESCRIPTION": df["_description"].apply(_clean),
    })

    out = out[out["TASK"].ne("")].reset_index(drop=True)
    return out


def exportar_excel(df: pd.DataFrame, output_path: str | Path | BytesIO) -> None:
    """Exporta con formato profesional, anchos ajustados y AC/WO combinados por grupo."""
    df = df[OUTPUT_COLUMNS].copy()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Programa")

    # Reabrir para aplicar formato.
    if isinstance(output_path, BytesIO):
        output_path.seek(0)
        wb = load_workbook(output_path)
    else:
        wb = load_workbook(output_path)

    ws = wb["Programa"]
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin_gray)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = {"A": 12, "B": 12, "C": 22, "D": 95}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row[0].row].height = 42

    # Combinar AC y WO por bloque AC+WO consecutivo.
    start = 2
    while start <= ws.max_row:
        ac = ws.cell(start, 1).value
        wo = ws.cell(start, 2).value
        end = start
        while end + 1 <= ws.max_row and ws.cell(end + 1, 1).value == ac and ws.cell(end + 1, 2).value == wo:
            end += 1
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
        ws.cell(start, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(start, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(start, 1).font = Font(bold=True)
        ws.cell(start, 2).font = Font(bold=True)
        start = end + 1

    # Tabla solo si no hay celdas combinadas? Excel no permite tabla sobre rangos con merges de forma confiable.
    # Por eso dejamos autofiltro simple en encabezados.
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    if isinstance(output_path, BytesIO):
        output_path.seek(0)
        wb.save(output_path)
        output_path.seek(0)
    else:
        wb.save(output_path)


def procesar_y_guardar(input_file, output_path: str | Path, incluir_defectos_sin_eo: bool = False) -> Path:
    df = procesar_programa(input_file, incluir_defectos_sin_eo=incluir_defectos_sin_eo)
    exportar_excel(df, output_path)
    return Path(output_path)
